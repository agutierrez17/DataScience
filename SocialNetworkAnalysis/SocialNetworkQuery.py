import pandas as pd # Dataframes
import matplotlib.pyplot as plt # Create plots, graphs, etc.
import pyodbc # Used for querying in data from an external database environment
import networkx as nx # Used for creating social netowrk diagrams
import mplcursors # Used for adding annotations to matplotlib figures

# Default to centering matplotlib within the notebook
from IPython.core.display import HTML
HTML("""
<style>
.output_png {
    display: table-cell;
    text-align: center;
    vertical-align: middle;
}
</style>
""")

# Ignore warnings
import warnings 
warnings.filterwarnings("ignore")

# Connect to SQL Cursor
conn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};'
                      'Trusted_Connection=yes;')

sql = """
SELECT DISTINCT
R.ID,
C.[Last Name],
C.[First Name],
C.Age,
C.City,
C.ST,
C.ZIP,
C.Alumnus,
C.School,
C.GradYear,
C.[Lifetime Giving],
[Related ID],
[Relationship Type],
[Status],
[Notes]
FROM [dbo].[ConstituentsView] C WITH (NOLOCK) 
INNER JOIN [dbo].[Relationships] R WITH (NOLOCK) ON C.ID = R.[Related ID]

WHERE
[Status] = 'Confirmed'
"""
df = pd.read_sql(sql,conn)

# Create graph
G = nx.Graph()

# Create nodes and edges
edges = df[['ID','Related ID']].values.tolist()

# Count how many times each pair attended together
edge_df = pd.DataFrame(edges, columns=['Source', 'Target'])
edge_weights = edge_df.value_counts().reset_index(name='weight')

# Add edges with weights
for _, row in edge_weights.iterrows():
    G.add_edge(row['Source'], row['Target'], weight=row['weight'])

# Populate diagram
pos = nx.spring_layout(G, seed=42)  # Layout for consistent visuals
weights = [G[u][v]['weight'] for u, v in G.edges()]

# Calculate centricity measures
degree_centrality = nx.degree_centrality(G)
betweenness_centrality = nx.betweenness_centrality(G, weight='weight')
closeness_centrality = nx.closeness_centrality(G)
eigenvector_centrality = nx.eigenvector_centrality(G, weight='weight', max_iter=10000)

# Create lists for node data
nodes = []
x_coords = []
y_coords = []
degrees = []
betweennesses = []
closenesses = []
eigenvectors = []

# Append values to the list
for node in G.nodes():
    nodes.append(node)
    x_coords.append(pos[node][0])
    y_coords.append(pos[node][1])
    degrees.append(degree_centrality[node])
    betweennesses.append(betweenness_centrality[node])
    closenesses.append(closeness_centrality[node])
    eigenvectors.append(eigenvector_centrality[node])

# Create DF for centricity meaures
rel_df = pd.DataFrame({
    'ID': nodes,
    'X': x_coords,
    'Y': y_coords,
    'Degree Centrality': degrees,
    'Betweenness Centrality': betweennesses,
    'Closeness Centrality': closenesses,
    'Eigenvector Centrality': eigenvectors
    })

rel_df = pd.merge(rel_df, df, how='left', on='ID')

# Create DF for edges
edge_weights = edge_weights.rename(columns={'Source': 'ID','Target': 'Related ID'})
edge_weights = pd.merge(edge_weights, df, how='left', on=['ID','Related ID'])
edge_weights = edge_weights.rename(columns={'ID': 'Source','Related ID': 'Target'})
edge_weights = edge_weights[['Source','Target','weight','Relationship Type','Status','Notes']]

# Write full data to Excel
rel_df.to_excel(writer,sheet_name='Nodes')
edge_weights.to_excel(writer,sheet_name='Edges')

# Close Excel Writer
writer.close()
print('All data written to Excel.')

from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
from openpyxl.utils import get_column_letter

# Open up workbook with OpenPyxl
print('Opening up Sample.xlsx Excel sheet with OpenPyxl...')
book = load_workbook(path)

# Create a currency style
currency_style = NamedStyle(name="currency_style", number_format='"$"#,##0')

# Loop through sheets, AutoFit all columns, delete first row, format currency fields
print('Looping through sheets, formatting columns...')
for sheet in book:
    sheet.delete_cols(1)

    # setting the column width
    for col in range(1, sheet.max_column + 1):
        max_length = 0
        column = get_column_letter(col)
        
        for row in sheet.iter_rows(min_col=col, max_col=col):
            for cell in row:
                try:
                    if len(str(cell.value)) > 90:
                        max_length = 90
                    elif len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass

        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column].width = adjusted_width

        # format as currency
        column_name = sheet.cell(row=1, column=col).value
        if 'Amount' in column_name or 'Value' in column_name or 'Revenue' in column_name or 'Balance' in column_name:
            for row in range(2, sheet.max_row + 1):
                cell = sheet.cell(row=row, column=col)
                cell.style = currency_style

print('Finished formatting workbook, closing file.')
book.save(path)   
print('')
