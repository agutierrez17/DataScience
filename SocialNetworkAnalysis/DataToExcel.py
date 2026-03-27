import pandas as pd
import pyodbc
import warnings
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
from openpyxl.utils import get_column_letter
import google.auth
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from googleapiclient.http import MediaFileUpload

warnings.filterwarnings("ignore")

# Connect to database and open SQL cursor
print('Connecting to database...')
print('')
conn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};'
                      'Trusted_Connection=yes;')
cursor = conn.cursor()

# Query Constituents view to start
print('Querying data from Constituents view...')
sql = """
SELECT
[ID],
[Last Name],
[First Name],
[Deceased],
[Birthdate],
[Age],
[Manager],
[Address Line 1],
[City],
[ST],
[ZIP],
[Country],
[Metro Area],
[Geo Area Type],
[ZIP Code Population],
[Metro Median Income],
[Metro Median Income (Married Families)],
[Metro % of Households Married Families],
[Alumnus],
[School],
[GradYear],
[Lifetime Giving],
[Last Gift Date],
[Last Gift Amount],
[Last Gift Type],
[Last Gift Fund],
[Events Attended],
[Last Event Date]
FROM [Philanthropy].[dbo].[ConstituentsView]
"""
df = pd.read_sql(sql,conn)

# Open up pandas ExcelWriter, write Constituents data to Excel sheet
print('')
print('Writing Constituents data to SampleData.xlsx Excel sheet...')
print('')
df.to_excel(writer,sheet_name='Constituents')

# Close Excel Writer
writer.close()
print('All data written to Excel.')
print('')

# Open up workbook with OpenPyxl
print('Opening up Sample.xlsx Excel sheet with OpenPyxl...')
book = load_workbook(path)

# Create a currency style
currency_style = NamedStyle(name="currency_style", number_format='"$"#,##0')

# Loop through sheets, AutoFit all columns, delete first row, format currency fields
print('Looping through sheets, formatting columns...')
for sheet in book:
    sheet.delete_cols(1)

    # setting the column width
    for col in range(1, sheet.max_column + 1):
        max_length = 0
        column = get_column_letter(col)
        
        for row in sheet.iter_rows(min_col=col, max_col=col):
            for cell in row:
                try:
                    if len(str(cell.value)) > 90:
                        max_length = 90
                    elif len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass

        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column].width = adjusted_width

        # format as currency
        column_name = sheet.cell(row=1, column=col).value
        if 'Amount' in column_name or 'Value' in column_name or 'Revenue' in column_name or 'Balance' in column_name:
            for row in range(2, sheet.max_row + 1):
                cell = sheet.cell(row=row, column=col)
                cell.style = currency_style

print('Finished formatting workbook, closing file.')
book.save(path)   
cursor.close()
